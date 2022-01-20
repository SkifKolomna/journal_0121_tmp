import datetime
import math
from datetime import datetime
from pprint import pprint

from dateutil.relativedelta import relativedelta
from django.contrib.auth.decorators import login_required
from django.db import connections
from django.http import HttpResponse
from django.shortcuts import redirect
from django.utils import timezone
from openpyxl import *
from openpyxl.styles import Font
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

from tasks.models import Task


@login_required
def export_report_reu_all(request):
    report_queryset_old = Task.objects.all()
    report_queryset_old = report_queryset_old.filter(reu__icontains='РЭУ')
    report_queryset_old = report_queryset_old.filter(created_on__date=datetime.now())

    report_reu = set()

    for x in report_queryset_old.distinct():
        report_reu.add(x.reu)
    report_difference = {'ЭМУ', 'Служба благоустройства', None, 'Аварийная служба', 'ЛифтМастер', 'Газовая служба',
                         'ВКХ Водоканал', 'ТК Тепло Коломны', 'Электросеть'}
    report_reu = sorted(report_reu.difference(report_difference))
    if report_reu:
        return report(report_reu, report_queryset_old, request)
    else:
        return redirect('tasks:tasks-list')


@login_required
def export_report_tools_all(request):
    report_queryset_old = Task.objects.all()
    report_queryset_old = report_queryset_old.exclude(reu__icontains='РЭУ')
    report_queryset_old = report_queryset_old.filter(created_on__date=datetime.now())

    report_reu = set()

    for x in report_queryset_old.distinct():
        report_reu.add(x.reu)
    report_difference = {None, }
    report_reu = sorted(report_reu.difference(report_difference))
    if report_reu:
        return report(report_reu, report_queryset_old, request)
    else:
        return redirect('tasks:tasks-list')


@login_required
def export_report_ac(request):
    report_queryset_old = Task.objects.all()
    report_queryset_old = report_queryset_old.filter(reu='Аварийная служба')
    now = datetime.now()
    today = datetime(now.year, now.month, now.day, 8, 00)
    then = today - relativedelta(days=1)
    report_queryset_old = report_queryset_old.filter(
        created_on__range=(
            timezone.make_aware(then),
            timezone.make_aware(today)
        )
    )
    report_reu = set()
    for x in report_queryset_old.distinct():
        report_reu.add(x.reu)
    report_difference = {None, }
    report_reu = sorted(report_reu.difference(report_difference))
    if report_reu:
        return report(report_reu, report_queryset_old, request)
    else:
        return redirect('tasks:tasks-list')


@login_required
def export_report_tools(request):
    report_queryset_old = Task.objects.all()
    report_queryset_old = report_queryset_old.exclude(status_task__icontains='выполнена')
    report_queryset_old = report_queryset_old.exclude(reu__icontains='РЭУ')

    report_reu = set()

    for x in report_queryset_old.distinct():
        report_reu.add(x.reu)
    report_difference = {None, }  # 'ЭМУ'
    report_reu = sorted(report_reu.difference(report_difference))
    return report(report_reu, report_queryset_old, request)


@login_required
def export_report_reu(request):
    report_queryset_old = Task.objects.all().exclude(status_task__icontains='выполнена')

    report_reu = set()
    for x in report_queryset_old.distinct():
        report_reu.add(x.reu)
    report_difference = {'ЭМУ', 'Служба благоустройства', None, '', 'Аварийная служба', 'ЛифтМастер', 'Газовая служба',
                         'ВКХ Водоканал', 'ТК Тепло Коломны', 'Электросеть'}
    report_reu = sorted(report_reu.difference(report_difference))
    return report(report_reu, report_queryset_old, request)


@login_required
def export_report_month(request):
    report_queryset_old = Task.objects.all()
    mon = timezone.now() + relativedelta(months=-1)
    report_queryset_old = report_queryset_old.filter(created_on__year=mon.year, created_on__month=mon.month)
    report_reu = set()
    for x in report_queryset_old.distinct():
        report_reu.add(x.reu)
    report_difference = {None, }
    report_reu = sorted(report_reu.difference(report_difference))
    return report(report_reu, report_queryset_old, request)


@login_required
def export_report_all(request):
    report_queryset_old = Task.objects.all()
    report_reu = set()
    # print(report_queryset_old.query)
    # print(report_queryset_old.query.__str__())
    for x in report_queryset_old.distinct():
        # print(x)
        report_reu.add(x.reu)
    report_difference = {None, }
    report_reu = sorted(report_reu.difference(report_difference))
    return report(report_reu, report_queryset_old, request)


@login_required
def export_report_month_tek(request):
    report_queryset_old = Task.objects.all()
    now = timezone.now()
    # report_queryset_old = report_queryset_old.filter(created_on__year=now.year,
    #                                                  created_on__month=now.month)
    report_queryset_old = report_queryset_old.filter(created_on__year=now.year)
    report_queryset_old = report_queryset_old.filter(created_on__month=now.month)
    report_reu = set()
    for x in report_queryset_old.distinct():
        report_reu.add(x.reu)
    report_difference = {None, }
    report_reu = sorted(report_reu.difference(report_difference))
    return report(report_reu, report_queryset_old, request)


@login_required
def edit_lift(request):
    all_task = Task.objects.all()
    task_lift = all_task.filter(reu__icontains='Мастер-лифт')
    for x in task_lift:
        print(x.id, x.reu)
        x.reu = 'ЛифтМастер'
        x.save()
    source_task_lift = all_task.filter(source_task__icontains='Мастер-лифт')
    for x in source_task_lift:
        print(x.id, x.source_task)
        x.source_task = 'ЛифтМастер'
        x.save()


def report(report_reu, report_queryset_old, request):
    workbook = Workbook()
    workbook.remove(workbook.active)
    for report_name in report_reu:
        if report_name:
            report_queryset = report_queryset_old.filter(reu__icontains=report_name)

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = 'attachment; filename={date} report.xlsx'.format(
                date=datetime.now().strftime('%Y.%m.%d %H:%M'),
            )
            worksheet = workbook.create_sheet(report_name)

            # Define the titles for columns
            columns = [
                'п/п',
                'Дата/время',
                '№',
                'Адрес заявки',
                'Телефон',
                'Содержание заявки',
                'Передано',
                'Статус',
                'Источник',
                'ФИО',
                'Комментарий',
            ]
            row_num = 1

            # Assign the titles for each cell of the header
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title
                cell.fill = PatternFill(fill_type='solid', start_color='ff8327', end_color='ff8327')
                # Данный код позволяет делать оформление цветом ячейки
                cell.font = Font(size=14, bold=True)
                cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

                border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom
                )
                side = Side(border_style='thin')
                border.left = side
                border.right = side
                border.top = side
                border.bottom = side

                cell.border = border

            # Iterate through all movies

            for task in report_queryset:
                row_num += 1
                row = [
                    '{}'.format(row_num - 1),
                    '{}'.format(timezone.make_naive(task.created_on).strftime('%d.%m.%y %H:%M')),
                    '{}'.format(task.id),
                    return_str_address(task),
                    return_str_not_null(task.phone),
                    task.description,
                    return_str_tranmission(task),
                    return_status_task(task),
                    task.source_task,
                    return_str_fio(task),
                    task.comment_status,
                ]

                for col_num, cell_value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = cell_value
                    cell.font = Font(size=12)
                    cell.border = border

                    if task.description:
                        if "повтор" in task.description.lower():
                            cell.fill = PatternFill(fill_type='solid', start_color='F4A460', end_color='F4A460')

                    if task.comment_status:
                        if "повтор" in task.comment_status.lower():
                            cell.fill = PatternFill(fill_type='solid', start_color='F4A460', end_color='F4A460')

                    if task.source_task:
                        if "администрация" in task.source_task:
                            cell.font = Font(size=12, color='FF0000', bold=True)

                    if col_num == 6 or col_num == 7 or col_num == 11:
                        worksheet.column_dimensions[get_column_letter(col_num)].width = 36
                        # cell.alignment = Alignment(wrapText=True)
                    # else:
                    #     worksheet.column_dimensions[get_column_letter(col_num)].width = 16

                    if col_num == 1:
                        worksheet.column_dimensions[get_column_letter(col_num)].width = 5
                    if col_num == 2:
                        worksheet.column_dimensions[get_column_letter(col_num)].width = 16
                    if col_num == 3:
                        worksheet.column_dimensions[get_column_letter(col_num)].width = 7
                    if col_num == 4:
                        worksheet.column_dimensions[get_column_letter(col_num)].width = 36
                        worksheet.column_dimensions[get_column_letter(col_num)].bestFit = True
                        # cell.alignment = Alignment(wrapText=True)
                    if col_num == 5:
                        worksheet.column_dimensions[get_column_letter(col_num)].width = 16
                        # cell.alignment = Alignment(wrapText=True)
                    if col_num == 8:
                        worksheet.column_dimensions[get_column_letter(col_num)].width = 16
                        worksheet.column_dimensions[get_column_letter(col_num)].bestFit = True
                        # cell.alignment = Alignment(wrapText=True)
                    if col_num == 9:
                        worksheet.column_dimensions[get_column_letter(col_num)].width = 14
                        worksheet.column_dimensions[get_column_letter(col_num)].bestFit = True
                        # cell.alignment = Alignment(wrapText=True)
                    if col_num == 10:
                        worksheet.column_dimensions[get_column_letter(col_num)].width = 12
                        worksheet.column_dimensions[get_column_letter(col_num)].bestFit = True
                    cell.alignment = Alignment(horizontal='left', vertical='top')

            col_width = []

            for i in range(len(next(worksheet.iter_rows()))):
                col_letter = get_column_letter(i + 1)
                minimum_width = 4
                current_width = worksheet.column_dimensions[col_letter].width
                if not current_width or current_width < minimum_width:
                    worksheet.column_dimensions[col_letter].width = minimum_width
                # worksheet.column_dimensions[col_letter].width = minimum_width

                col_width.append(worksheet.column_dimensions[col_letter].width)

            for i, row in enumerate(worksheet):
                default_height = 6  # 10.5  # Corresponding to font size 12 12,5

                multiples_of_font_size = [default_height]
                for j, cell in enumerate(row):
                    wrap_text = True
                    vertical = "top"
                    if cell.value is not None:
                        mul = 0
                        for v in str(cell.value).split('\n'):
                            mul += math.ceil(len(v) / col_width[j]) * (cell.font.size + 8)  # добавил 5 было 3

                        if mul > 0:
                            multiples_of_font_size.append(mul)

                    cell.alignment = Alignment(wrap_text=wrap_text, vertical=vertical)
                    if i == 0:
                        cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

                    original_height = worksheet.row_dimensions[i + 1].height
                    if original_height is None:
                        original_height = default_height

                    new_height = max(multiples_of_font_size)
                    if original_height < new_height:
                        worksheet.row_dimensions[i + 1].height = new_height

            # Printer Settings
            worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
            worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
            worksheet.freeze_panes = "A2"
            worksheet.print_title_rows = '1:1'

            wsprops = worksheet.sheet_properties

            # wsprops.tabColor = "1072BA"
            wsprops.filterMode = False

            # wsprops.pageSetUpPr = PageSetupProperties(fitToPage=True) #, autoPageBreaks=False)
            wsprops.pageSetUpPr = PageSetupProperties(fitToPage=True)  # , autoPageBreaks=False)

            worksheet.page_setup.fitToHeight = False
            worksheet.page_setup.fitToWidth = True

            wsmargin = worksheet.page_margins
            wsmargin.left = 1 / 3  # (left=0.75, right=0.75, top=1, bottom=1, header=0.5, footer=0.5)
            wsmargin.right = 1 / 4  # (left=0.75, right=0.75, top=1, bottom=1, header=0.5, footer=0.5)
            wsmargin.top = 1 / 3  # (left=0.75, right=0.75, top=1, bottom=1, header=0.5, footer=0.5)
            wsmargin.bottom = 1 / 4  # (left=0.75, right=0.75, top=1, bottom=1, header=0.5, footer=0.5)

            wsprops.outlinePr.summaryBelow = False
            wsprops.outlinePr.applyStyles = True

            workbook.save(response)

    return response


def return_str_address(self):
    str_address = ""
    if self.address:
        str_address += str(self.address)
    if self.apartment:
        if str_address:
            str_address += ", кв." + self.apartment
    if self.porch:
        if str_address:
            str_address += ", п." + str(self.porch)
    if self.floor:
        if str_address:
            str_address += ", эт." + str(self.floor)
    return str_address


def return_str_not_null(self):
    str_not_null = ""
    if self is not None:
        str_not_null = str(self)
    return str_not_null


def return_str_tranmission(self):
    str_tranmission = ""
    if self.reu:
        str_tranmission += self.reu
    if self.executor:
        str_tranmission += ", " + self.executor
    if self.transmission_time:
        str_tranmission += ", в " + str(self.transmission_time.strftime('%H:%M'))
    return str_tranmission


def return_status_task(self):
    str_status_task = ""
    if self.status_task == "выполнена":
        # print('выполнена')
        str_status_task += self.status_task
        if self.status_time is not None:
            str_status_task += " " + str(timezone.make_naive(self.status_time).strftime('%d.%m.%Y %H:%M'))
    else:
        str_status_task = self.status_task
    return str_status_task


def return_str_fio(self):
    str_fio = ""
    if self.surname_name:
        str_fio += self.surname_name
    if self.first_name:
        str_fio += " " + self.first_name
    if self.patronymic_name:
        str_fio += " " + self.patronymic_name
    if self.author_eds:
        str_fio += " " + self.author_eds

    return str_fio
